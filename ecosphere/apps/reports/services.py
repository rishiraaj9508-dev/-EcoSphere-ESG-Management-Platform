import csv
import io
from django.http import HttpResponse
from django.utils import timezone

# 3rd party exports
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# Models
from apps.environmental.models import CarbonEmission, SustainabilityGoal
from apps.social.models import CSRParticipation, TrainingCompletion, DiversityMetric
from apps.governance.models import PolicyAcknowledgement, Audit, ComplianceIssue
from apps.dashboard.models import DepartmentESGScore

class ReportService:
    @staticmethod
    def build_environmental(date_from, date_to, depts):
        emissions = CarbonEmission.objects.filter(reporting_period__range=[date_from, date_to])
        goals = SustainabilityGoal.objects.filter(deadline__range=[date_from, date_to])
        if depts:
            emissions = emissions.filter(department__in=depts)
            goals = goals.filter(department__in=depts)
        
        headers = ["Date/Deadline", "Department", "Entity Type", "Metric / Title", "Value / Target", "Status"]
        rows = []
        for e in emissions:
            rows.append([
                e.reporting_period.strftime("%Y-%m-%d"),
                e.department.name,
                "Carbon Emission",
                e.activity_type,
                f"{e.co2e_value} mt CO2e",
                "Calculated"
            ])
        for g in goals:
            rows.append([
                g.deadline.strftime("%Y-%m-%d"),
                g.department.name,
                "Sustainability Goal",
                g.title,
                f"{g.current_value} / {g.target_value}",
                g.status
            ])
        return {"title": "Environmental Performance Report", "headers": headers, "rows": rows}

    @staticmethod
    def build_social(date_from, date_to, depts):
        csr = CSRParticipation.objects.filter(submitted_at__date__range=[date_from, date_to])
        trainings = TrainingCompletion.objects.filter(training__training_date__range=[date_from, date_to])
        diversity = DiversityMetric.objects.filter(reporting_period__range=[date_from, date_to])
        if depts:
            csr = csr.filter(activity__department__in=depts)
            trainings = trainings.filter(training__department__in=depts)
            diversity = diversity.filter(department__in=depts)

        headers = ["Date", "Department", "Category", "Participant", "Details / Metric", "Status / Value"]
        rows = []
        for c in csr:
            rows.append([
                c.submitted_at.strftime("%Y-%m-%d") if c.submitted_at else "N/A",
                c.activity.department.name,
                "CSR Activity",
                c.employee.username,
                c.activity.title,
                c.status
            ])
        for t in trainings:
            rows.append([
                t.training.training_date.strftime("%Y-%m-%d"),
                t.training.department.name,
                "Training Course",
                t.employee.username,
                t.training.title,
                "Completed" if t.completed else "Pending"
            ])
        for d in diversity:
            rows.append([
                d.reporting_period.strftime("%Y-%m-%d"),
                d.department.name,
                "Diversity Metric",
                "N/A",
                d.metric_type,
                f"{d.value} {d.unit}"
            ])
        return {"title": "Social & Labor Performance Report", "headers": headers, "rows": rows}

    @staticmethod
    def build_governance(date_from, date_to, depts):
        # Filter audits and compliance issues
        audits = Audit.objects.filter(audit_date__range=[date_from, date_to])
        issues = ComplianceIssue.objects.filter(created_at__date__range=[date_from, date_to])
        if depts:
            audits = audits.filter(department__in=depts)
            issues = issues.filter(department__in=depts)

        headers = ["Date", "Department", "Entity Type", "Auditor / Owner", "Title / Scope", "Status / Severity"]
        rows = []
        for a in audits:
            rows.append([
                a.audit_date.strftime("%Y-%m-%d"),
                a.department.name,
                "Audit Record",
                a.auditor,
                a.title,
                a.status
            ])
        for i in issues:
            rows.append([
                i.created_at.strftime("%Y-%m-%d"),
                i.department.name,
                "Compliance Issue",
                i.owner.username,
                i.title,
                f"{i.status} ({i.severity})"
            ])
        return {"title": "Governance & Compliance Report", "headers": headers, "rows": rows}

    @staticmethod
    def build_esg_summary(date_from, date_to, depts):
        # Snapshot score averages for selected departments
        scores = DepartmentESGScore.objects.all()
        if depts:
            scores = scores.filter(department__in=depts)
        
        headers = ["Department", "Environmental Score", "Social Score", "Governance Score", "Overall ESG Score", "Last Recalculated"]
        rows = []
        for s in scores:
            rows.append([
                s.department.name,
                f"{s.environmental_score}/100",
                f"{s.social_score}/100",
                f"{s.governance_score}/100",
                f"{s.overall_score}/100",
                s.last_calculated_at.strftime("%Y-%m-%d %H:%M") if s.last_calculated_at else "N/A"
            ])
        return {"title": "ESG Overall Summary Disclosures", "headers": headers, "rows": rows}

    # Format Exporters
    @staticmethod
    def export_csv(filename, report_data):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([report_data["title"]])
        writer.writerow([])
        writer.writerow(report_data["headers"])
        for row in report_data["rows"]:
            writer.writerow(row)
        return response

    @staticmethod
    def export_excel(filename, report_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ESG Report"

        # Apply basic theme colors
        title_font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        regular_font = Font(name='Segoe UI', size=11)
        
        dark_blue_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
        green_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')

        # Title Block
        ws.merge_cells('A1:F1')
        ws['A1'] = report_data["title"]
        ws['A1'].font = title_font
        ws['A1'].fill = green_fill
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 40

        # Headers
        ws.append([]) # spacer
        ws.append(report_data["headers"])
        header_row_idx = 3
        ws.row_dimensions[header_row_idx].height = 25
        for col_idx in range(1, len(report_data["headers"]) + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.font = header_font
            cell.fill = dark_blue_fill
            cell.alignment = Alignment(vertical='center', indent=1)

        # Rows
        for row in report_data["rows"]:
            ws.append(row)
            curr_row_idx = ws.max_row
            ws.row_dimensions[curr_row_idx].height = 20
            for col_idx in range(1, len(row) + 1):
                cell = ws.cell(row=curr_row_idx, column=col_idx)
                cell.font = regular_font
                cell.alignment = Alignment(vertical='center', indent=1)

        # Autofit Columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response

    @staticmethod
    def export_pdf(filename, title, report_data):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'

        doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        story = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=18,
            textColor=colors.HexColor('#10B981'),
            spaceAfter=15
        )
        body_style = ParagraphStyle(
            'ReportBody',
            parent=styles['BodyText'],
            fontName='Helvetica',
            fontSize=9,
            textColor=colors.HexColor('#374151')
        )
        header_style = ParagraphStyle(
            'ReportHeader',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10,
            textColor=colors.white
        )

        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 10))

        # Format Table Data
        table_data = []
        table_data.append([Paragraph(h, header_style) for h in report_data["headers"]])
        
        for r in report_data["rows"]:
            table_data.append([Paragraph(str(cell), body_style) for cell in r])

        col_widths = [75, 80, 85, 95, 115, 75] # Adjust width dynamically to fit letter page
        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F9FAFB'), colors.white]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
        ]))
        
        story.append(t)
        doc.build(story)
        return response
